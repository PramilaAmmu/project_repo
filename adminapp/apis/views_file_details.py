from django.shortcuts import render , redirect
from adminapp.models import uploadedFileDetails
from adminapp.serializers import uploadedFileDetailsSerializer
from django.contrib import messages
from rest_framework.views import APIView
from datetime import datetime
import openpyxl


class UploadFileAPI(APIView):
    def get(self,request):
        try:
        	if request.session.has_key("user_id") & request.session.has_key('role_type') & request.session.has_key('login_status'):
        		if request.session['role_type'] != 3:
        			output_data = {}
        			user_id = request.session["user_id"]

        			'''fetching all file details added by the logged in user '''
        			file_details_qs = uploadedFileDetails.objects.filter(user_id = user_id)
        			serialized_var = uploadedFileDetailsSerializer(file_details_qs,many=True)

        			output_data['Status'] = "Success"
        			output_data['Message'] = "Successfully fetched uploaded file details"
        			output_data['file_details'] = serialized_var.data

        			return render(request,'upload_file_page.html',output_data)

        		else:
        			return redirect('/login/')
        	else:
        		return redirect('/')

        except Exception:
            return redirect('/')

    def post(self,request):
        try:
            if request.session.has_key("user_id") & request.session.has_key('role_type') & request.session.has_key('login_status') & request.session.has_key('user_name'):
            	
                output_data = {}
                today_date  = datetime.now().strftime('%Y-%m-%d')
                user_id = request.session["user_id"]

                try:
                    excel_file = request.FILES["excel_file"]
                    wb = openpyxl.load_workbook(excel_file)
				    # getting active sheet
                    worksheet = wb.active
                    first_row = list(worksheet.rows)[0]

                    if first_row[0].value == "sl_no" and first_row[1].value == "name" and first_row[2].value == "education" and first_row[3].value == "city" and first_row[4].value == "email_id":
                        all_file_details_list = []
                        for row in worksheet.iter_rows(2,worksheet.max_row):
                            individual_file_data_insert = {}
                            # individual_file_data_insert['sl_no'] = row[0].value
                            individual_file_data_insert['name'] = row[1].value
                            individual_file_data_insert['education'] = row[2].value
                            individual_file_data_insert['city'] = row[3].value
                            individual_file_data_insert['email_id'] = row[4].value
                            individual_file_data_insert['user_id'] = user_id
                            individual_file_data_insert['added_by'] = request.session["user_name"]
                            individual_file_data_insert['added_date'] = today_date
                            all_file_details_list.append(individual_file_data_insert)

                        try:
                            serialized_data_var = uploadedFileDetailsSerializer(data = all_file_details_list,many=True)
                            if serialized_data_var.is_valid(raise_exception = True):
                                serialized_data_var.save()
                                output_data['Status'] = 'Success'
                                output_data['Message'] = 'Data added successfully'
                                messages.success(request, output_data["Message"])
                                return redirect('/upload_file/')
                        except Exception:
                            output_data['Status'] = 'Failure'
                            output_data['Message'] = 'Data could not be inserted successfully'
                            messages.warning(request, output_data["Message"])
                            return redirect('/upload_file/')
                    else:
                        output_data['Status'] = "Failure"
                        output_data['Message'] = "Invalid file uploaded. Fields names not matching with what is provided in sample file"
                        messages.warning(request, output_data["Message"])
                        return redirect('/upload_file/')

                except Exception:
                    output_data['Status'] = "Failure"
                    output_data['Message'] ="Invalid File Upload"
                    messages.warning(request, output_data['Message'])
                    return redirect('/upload_file/')
                    
            else:
            	return redirect('/')

        except Exception:
            return redirect('/')